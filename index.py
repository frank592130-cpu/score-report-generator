from flask import Flask, request, send_file, render_template_string, jsonify
import math
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ════════════════════════════════
# 核心 Excel 繪製邏輯
# ════════════════════════════════
FONT_NAME = "新細明體"
FILLS = {
    "A++": None,
    "A+":  PatternFill("solid", fgColor="E6E6E6"),
    "A":   PatternFill("solid", fgColor="BFBFBF"),
    "B++": PatternFill("solid", fgColor="808080"),
    "":    PatternFill("solid", fgColor="808080"),
}

def _med(): return Side(style="medium", color="000000")
def _thn(): return Side(style="thin",   color="000000")
def all_thin():
    s = _thn()
    return Border(left=s, right=s, top=s, bottom=s)

def outer_med(r, c, r1, c1, r2, c2):
    return Border(
        left   = _med() if c == c1 else _thn(),
        right  = _med() if c == c2 else _thn(),
        top    = _med() if r == r1 else _thn(),
        bottom = _med() if r == r2 else _thn(),
    )

def sc(ws, row, col, value, bold=False, size=10, fill=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name=FONT_NAME, bold=bold, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if border: cell.border = border
    if fill: cell.fill = fill
    return cell

# ---------------------------------------------------------
# 1. 讀取邏輯：抓取 P 欄 (Index 15)
# ---------------------------------------------------------
def read_students_initial(file_stream):
    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        ws = wb.active
        students = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 16: continue
            name = str(row[14]).strip() if row[14] is not None else "" 
            student_id = str(row[4]).strip() if row[4] is not None else "" 
            if name in ["", "預設標準答案", "None"]: continue
            try:
                x_val = float(row[15]) if row[15] is not None else 0.0
                students.append({"id": student_id, "name": name, "x": x_val, "y": 0})
            except (ValueError, TypeError): continue
        return students
    except Exception as e:
        return []

# ---------------------------------------------------------
# 2. 生成 APP 貼上表邏輯
# ---------------------------------------------------------
@app.route('/generate_copy_list', methods=['POST'])
def generate_copy_list():
    students_json = request.form.get('students_json', '[]')
    ordered_names_raw = request.form.get('ordered_names', '')
    students = json.loads(students_json)
    student_map = {s['name']: s for s in students}
    ordered_names = [n.strip() for n in ordered_names_raw.split('\n') if n.strip()]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "補習班貼上專用"
    ws.cell(row=1, column=1, value="APP名單順序")
    ws.cell(row=1, column=2, value="總分 (表現)")
    
    for i, target_name in enumerate(ordered_names, start=2):
        ws.cell(row=i, column=1, value=target_name)
        if target_name in student_map:
            s = student_map[target_name]
            total = (float(s.get('x', 0)) / 25.0) * 85.0 + (float(s.get('y', 0)) / 6.0) * 15.0
            ws.cell(row=i, column=2, value=round(total, 2) if total > 0 else "")
        else:
            ws.cell(row=i, column=2, value="")
            
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="App_Score_Import.xlsx")

# ---------------------------------------------------------
# 3. 正式成績報表 (優化欄位寬度)
# ---------------------------------------------------------
def build_excel(students_data, exam_lines, ths):
    students_for_render = []
    for s in students_data:
        x, y = float(s.get('x', 0)), float(s.get('y', 0))
        total = (x / 25.0) * 85.0 + (y / 6.0) * 15.0
        students_for_render.append((s.get('name', ''), x, y, round(total, 2)))

    sorted_s = sorted(students_for_render, key=lambda x: -x[3])
    n = len(sorted_s)
    counts = {"A++": 0, "A+": 0, "A": 0, "B++": 0}
    for s in sorted_s:
        t = s[3]
        if t >= ths['th_app']: counts["A++"] += 1
        elif t >= ths['th_ap']: counts["A+"] += 1
        elif t >= ths['th_a']: counts["A"] += 1
        elif t >= ths['th_bpp']: counts["B++"] += 1

    rows_per_block = math.ceil((n + 2) / 3) if n > 0 else 1
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "成績報表"

    # --- 欄位寬度優化處 ---
    for b in range(3):
        base = b * 4 + 1
        ws.column_dimensions[get_column_letter(base)].width = 9       # 姓名 (A, E, I)
        ws.column_dimensions[get_column_letter(base + 1)].width = 5   # 選擇 (B, F, J) 縮短
        ws.column_dimensions[get_column_letter(base + 2)].width = 5   # 非選 (C, G, K) 縮短
        ws.column_dimensions[get_column_letter(base + 3)].width = 5.5 # 總分 (D, H, L) 縮短
    
    ws.column_dimensions["M"].width = 0.4
    for col_let in ["N", "O", "P"]: ws.column_dimensions[col_let].width = 7

    for b in range(3):
        base = b * 4 + 1
        for i, h in enumerate(["姓名", "選擇", "非選", "總分"]):
            sc(ws, 1, base + i, h, border=all_thin())

    for idx, (name, sel, nonsel, total) in enumerate(sorted_s):
        b, r = idx // rows_per_block, 2 + (idx % rows_per_block)
        col = b * 4 + 1
        g = ""
        if total >= ths['th_app']: g = "A++"
        elif total >= ths['th_ap']: g = "A+"
        elif total >= ths['th_a']: g = "A"
        elif total >= ths['th_bpp']: g = "B++"
        f = FILLS.get(g)
        for i, val in enumerate([name, sel, nonsel, total]):
            sc(ws, r, col + i, val, fill=f, border=all_thin())

    if n > 0:
        avg_pos = n
        b_avg, r_avg_start = avg_pos // rows_per_block, 2 + (avg_pos % rows_per_block)
        col_avg_base = b_avg * 4 + 1
        avg_vals = ["平均", round(sum(s[1] for s in students_for_render)/n, 2), 
                    round(sum(s[2] for s in students_for_render)/n, 2), 
                    round(sum(s[3] for s in students_for_render)/n, 2)]
        final_row = 2 + rows_per_block - 1
        for i, val in enumerate(avg_vals):
            curr_col = col_avg_base + i
            for fill_r in range(r_avg_start, final_row + 1):
                sc(ws, fill_r, curr_col, "", border=all_thin())
            if final_row > r_avg_start:
                ws.merge_cells(start_row=r_avg_start, start_column=curr_col, end_row=final_row, end_column=curr_col)
            sc(ws, r_avg_start, curr_col, val, bold=True, border=all_thin())

    TITLE_R1, TITLE_R2, TITLE_C1, TITLE_C2 = 2, 7, 14, 16
    ws.merge_cells(start_row=TITLE_R1, start_column=TITLE_C1, end_row=TITLE_R2, end_column=TITLE_C2)
    tc = ws.cell(row=TITLE_R1, column=TITLE_C1)
    tc.value = "\n".join(exam_lines)
    tc.font, tc.alignment = Font(name=FONT_NAME, bold=True, size=18), Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(TITLE_R1, TITLE_R2 + 1):
        for c in range(TITLE_C1, TITLE_C2 + 1):
            ws.cell(row=r, column=c).border = outer_med(r, c, TITLE_R1, TITLE_C1, TITLE_R2, TITLE_C2)

    visible_grades = [("A++", counts["A++"]), ("A+", counts["A+"]), ("A", counts["A"]), ("B++", counts["B++"])]
    for i, (g, cnt) in enumerate(visible_grades):
        row, f = TITLE_R2 + 4 + i, FILLS.get(g)
        for col, val in [(14, g), (15, cnt)]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font, cell.alignment = Font(name=FONT_NAME, bold=True, size=11), Alignment(horizontal="center", vertical="center")
            cell.border = outer_med(row, col, TITLE_R2+4, 14, TITLE_R2+7, 15)
            if f: cell.fill = f

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# UI 模板 (維持不變)
HTML_TEMPLATE = '''... (略，與前次相同) ...'''

@app.route('/')
def index(): return render_template_string(HTML_TEMPLATE)

@app.route('/upload_read', methods=['POST'])
def upload_read():
    file = request.files.get('file')
    if not file: return jsonify({"students": []})
    return jsonify({"students": read_students_initial(io.BytesIO(file.read()))})

@app.route('/analyze_full', methods=['POST'])
def analyze_full():
    data = request.get_json() or {}
    students = data.get('students', [])
    ths = {k: float(data.get(k, 0)) for k in ['th_app', 'th_ap', 'th_a', 'th_bpp']}
    counts = {"A++": 0, "A+": 0, "A": 0, "B++": 0}
    for s in students:
        total = (float(s.get('x', 0)) / 25.0) * 85.0 + (float(s.get('y', 0)) / 6.0) * 15.0
        if total >= ths['th_app']: counts["A++"] += 1
        elif total >= ths['th_ap']: counts["A+"] += 1
        elif total >= ths['th_a']: counts["A"] += 1
        elif total >= ths['th_bpp']: counts["B++"] += 1
    return jsonify({"counts": counts})

@app.route('/generate', methods=['POST'])
def generate():
    exam_name = request.form.get('exam_name', '成績報表')
    ths = {k: float(request.form.get(k, 0)) for k in ['th_app', 'th_ap', 'th_a', 'th_bpp']}
    students = json.loads(request.form.get('students_json', '[]'))
    parts = exam_name.strip().split()
    lines = [parts[0], " ".join(parts[1:-1]), parts[-1]] if len(parts) >= 3 else ["", exam_name, ""]
    buf = build_excel(students, lines, ths)
    return send_file(buf, as_attachment=True, download_name=f"{exam_name}.xlsx")

if __name__ == "__main__":
    app.run(debug=True)
