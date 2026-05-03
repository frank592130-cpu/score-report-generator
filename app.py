import math
import io
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="成績報表產生器", page_icon="📊", layout="centered")

# ════════════════════════════════
#  深色 / 淺色模式切換
# ════════════════════════════════
if "dark_mode" not in st.session_state:
    st.session_state.dark_mode = False

# ════════════════════════════════
#  全域 CSS 注入（漸層版）
# ════════════════════════════════
def inject_css(dark: bool):
    if dark:
        page_bg     = (
            "radial-gradient(ellipse 60% 50% at 15% 20%, rgba(79,142,247,0.13) 0%, transparent 70%),"
            "radial-gradient(ellipse 50% 45% at 85% 75%, rgba(167,139,250,0.11) 0%, transparent 70%),"
            "radial-gradient(ellipse 40% 35% at 50% 50%, rgba(56,189,248,0.06) 0%, transparent 60%),"
            "linear-gradient(145deg, #0a0e1a 0%, #0f1525 40%, #0d1a2e 70%, #0a1020 100%)"
        )
        surface     = "#151c30"
        surface2    = "#1a2240"
        border_col  = "rgba(79,142,247,0.22)"
        text        = "#dce8ff"
        text_sub    = "#7a8fb8"
        accent_a    = "#4f8ef7"
        accent_b    = "#a78bfa"
        accent_grad = "linear-gradient(135deg, #4f8ef7 0%, #a78bfa 100%)"
        accent_glow = "rgba(79,142,247,0.22)"
        accent_glow2= "rgba(167,139,250,0.18)"
        btn_bg      = "rgba(79,142,247,0.10)"
        btn_hover   = "rgba(79,142,247,0.20)"
        input_bg    = "#0f1525"
        success_bg  = "rgba(52,211,153,0.10)"
        success_col = "#34d399"
        err_bg      = "rgba(248,113,113,0.10)"
        err_col     = "#f87171"
        toggle_icon = "☀️"
        toggle_tip  = "切換為淺色模式"
        shadow_card = "0 8px 40px rgba(0,0,0,0.50), inset 0 1px 0 rgba(255,255,255,0.04)"
        shadow_btn  = "0 4px 20px rgba(79,142,247,0.35)"
        divider_grad= "linear-gradient(90deg, transparent, rgba(79,142,247,0.4), rgba(167,139,250,0.4), transparent)"
        title_grad  = "linear-gradient(135deg, #7cb8ff 0%, #c4b5fd 60%, #93c5fd 100%)"
        exp_hdr     = "#131a30"
        exp_body    = "#0f1525"
        upload_bg   = "#131a30"
        upload_hover= "rgba(79,142,247,0.08)"
    else:
        page_bg     = (
            "radial-gradient(ellipse 55% 45% at 10% 15%, rgba(59,110,240,0.10) 0%, transparent 65%),"
            "radial-gradient(ellipse 50% 40% at 90% 80%, rgba(124,58,237,0.09) 0%, transparent 65%),"
            "radial-gradient(ellipse 35% 30% at 55% 45%, rgba(14,165,233,0.07) 0%, transparent 55%),"
            "linear-gradient(150deg, #eef4ff 0%, #f5f0ff 35%, #e8f4fd 65%, #f0f8ff 100%)"
        )
        surface     = "#ffffff"
        surface2    = "#f0f5ff"
        border_col  = "rgba(99,128,220,0.22)"
        text        = "#1a1f3c"
        text_sub    = "#5a6490"
        accent_a    = "#3b6ef0"
        accent_b    = "#7c3aed"
        accent_grad = "linear-gradient(135deg, #3b6ef0 0%, #7c3aed 100%)"
        accent_glow = "rgba(59,110,240,0.16)"
        accent_glow2= "rgba(124,58,237,0.12)"
        btn_bg      = "rgba(59,110,240,0.07)"
        btn_hover   = "rgba(59,110,240,0.14)"
        input_bg    = "#f8fbff"
        success_bg  = "rgba(16,185,129,0.08)"
        success_col = "#059669"
        err_bg      = "rgba(220,38,38,0.07)"
        err_col     = "#dc2626"
        toggle_icon = "🌙"
        toggle_tip  = "切換為深色模式"
        shadow_card = "0 4px 32px rgba(59,110,240,0.10), 0 1px 4px rgba(0,0,0,0.04)"
        shadow_btn  = "0 4px 16px rgba(59,110,240,0.28)"
        divider_grad= "linear-gradient(90deg, transparent, rgba(59,110,240,0.3), rgba(124,58,237,0.3), transparent)"
        title_grad  = "linear-gradient(135deg, #2563eb 0%, #7c3aed 60%, #0ea5e9 100%)"
        exp_hdr     = "#f0f5ff"
        exp_body    = "#fafcff"
        upload_bg   = "#f0f5ff"
        upload_hover= "rgba(59,110,240,0.06)"

    css = f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

    /* ── Base & Background ── */
    html, body {{ margin: 0; padding: 0; }}
    .stApp {{
        background: {page_bg} !important;
        background-attachment: fixed !important;
        color: {text} !important;
        font-family: 'Noto Sans TC', sans-serif !important;
    }}
    .block-container {{
        padding-top: 2.2rem !important;
        padding-bottom: 3.5rem !important;
        max-width: 780px !important;
    }}
    * {{ box-sizing: border-box; }}

    /* ── 標題漸層文字 ── */
    .app-title {{
        font-size: 1.85rem;
        font-weight: 800;
        letter-spacing: -0.04em;
        display: flex;
        align-items: center;
        gap: 0.65rem;
        line-height: 1.1;
    }}
    .app-title .title-text {{
        background: {title_grad};
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }}
    .app-title .icon-badge {{
        background: {accent_grad};
        color: #fff;
        border-radius: 12px;
        width: 42px; height: 42px;
        display: inline-flex; align-items: center; justify-content: center;
        font-size: 1.15rem;
        box-shadow: 0 3px 16px {accent_glow}, 0 1px 4px rgba(0,0,0,0.15);
        flex-shrink: 0;
    }}
    .app-subtitle {{
        color: {text_sub};
        font-size: 0.84rem;
        margin-top: 0.3rem;
        margin-bottom: 1.6rem;
        letter-spacing: 0.01em;
    }}

    /* ── 漸層分隔線 ── */
    .styled-divider {{
        height: 1.5px;
        background: {divider_grad};
        margin: 1.4rem 0;
        border: none;
    }}

    /* ── Expander（卡片）── */
    [data-testid="stExpander"] {{
        background: {surface} !important;
        border: 1px solid {border_col} !important;
        border-radius: 18px !important;
        box-shadow: {shadow_card} !important;
        overflow: hidden !important;
    }}
    [data-testid="stExpander"] summary {{
        color: {text} !important;
        font-weight: 700 !important;
        font-size: 0.92rem !important;
        padding: 1rem 1.4rem !important;
        background: {exp_hdr} !important;
        letter-spacing: 0.01em !important;
    }}
    [data-testid="stExpander"] summary:hover {{
        filter: brightness(0.97) !important;
    }}
    [data-testid="stExpander"] > div > div {{
        padding: 1.3rem 1.4rem 1.5rem !important;
        background: {exp_body} !important;
    }}

    /* ── Input ── */
    .stTextInput input,
    .stNumberInput input {{
        background: {input_bg} !important;
        border: 1.5px solid {border_col} !important;
        border-radius: 11px !important;
        color: {text} !important;
        font-family: 'Noto Sans TC', sans-serif !important;
        font-size: 0.92rem !important;
        padding: 0.6rem 0.9rem !important;
        transition: border-color 0.25s, box-shadow 0.25s !important;
    }}
    .stTextInput input:focus,
    .stNumberInput input:focus {{
        border-color: {accent_a} !important;
        box-shadow: 0 0 0 3px {accent_glow}, 0 0 0 6px {accent_glow2} !important;
        outline: none !important;
    }}
    .stTextInput label,
    .stNumberInput label {{
        color: {text_sub} !important;
        font-size: 0.78rem !important;
        font-weight: 600 !important;
        letter-spacing: 0.04em !important;
    }}
    /* NumberInput +/- 按鈕 */
    .stNumberInput [data-testid="stNumberInputStepDown"],
    .stNumberInput [data-testid="stNumberInputStepUp"],
    .stNumberInput button {{
        background: {btn_bg} !important;
        color: {accent_a} !important;
        border: 1.5px solid {border_col} !important;
        border-radius: 8px !important;
    }}
    .stNumberInput [data-testid="stNumberInputStepDown"]:hover,
    .stNumberInput [data-testid="stNumberInputStepUp"]:hover,
    .stNumberInput button:hover {{
        background: {btn_hover} !important;
        border-color: {accent_a} !important;
    }}

    /* ── 一般按鈕 ── */
    .stButton > button {{
        background: {btn_bg} !important;
        color: {accent_a} !important;
        border: 1.5px solid {border_col} !important;
        border-radius: 11px !important;
        font-family: 'Noto Sans TC', sans-serif !important;
        font-weight: 600 !important;
        font-size: 0.9rem !important;
        padding: 0.55rem 1.5rem !important;
        transition: all 0.22s !important;
    }}
    .stButton > button:hover {{
        background: {btn_hover} !important;
        border-color: {accent_a} !important;
        transform: translateY(-1px) !important;
        box-shadow: 0 4px 18px {accent_glow} !important;
    }}
    /* ── 主要按鈕（漸層）── */
    .stButton > button[kind="primary"] {{
        background: {accent_grad} !important;
        color: #ffffff !important;
        border: none !important;
        box-shadow: {shadow_btn} !important;
    }}
    .stButton > button[kind="primary"]:hover {{
        filter: brightness(1.08) saturate(1.1) !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 28px {accent_glow}, 0 2px 8px {accent_glow2} !important;
    }}
    /* ── 下載按鈕（漸層）── */
    .stDownloadButton > button {{
        background: {accent_grad} !important;
        color: #fff !important;
        border: none !important;
        border-radius: 12px !important;
        font-family: 'Noto Sans TC', sans-serif !important;
        font-weight: 700 !important;
        font-size: 0.95rem !important;
        padding: 0.7rem 1.5rem !important;
        transition: all 0.22s !important;
        box-shadow: {shadow_btn} !important;
        letter-spacing: 0.02em !important;
    }}
    .stDownloadButton > button:hover {{
        filter: brightness(1.08) saturate(1.1) !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 10px 32px {accent_glow}, 0 3px 10px {accent_glow2} !important;
    }}

    /* ── 上傳區 ── */
    [data-testid="stFileUploaderDropzone"] {{
        background: {upload_bg} !important;
        border: 2px dashed {border_col} !important;
        border-radius: 16px !important;
        padding: 2.2rem !important;
        transition: border-color 0.25s, background 0.25s, box-shadow 0.25s !important;
    }}
    [data-testid="stFileUploaderDropzone"]:hover {{
        border-color: {accent_a} !important;
        background: {upload_hover} !important;
        box-shadow: 0 0 0 3px {accent_glow} !important;
    }}
    [data-testid="stFileUploaderDropzone"] * {{
        color: {text_sub} !important;
    }}
    /* 上傳按鈕本身 */
    [data-testid="stFileUploaderDropzone"] button,
    [data-testid="stFileUploaderDropzoneInstructions"] + div button {{
        background: {btn_bg} !important;
        color: {accent_a} !important;
        border: 1.5px solid {border_col} !important;
        border-radius: 9px !important;
    }}

    /* ── Alert ── */
    [data-testid="stAlert"] {{
        border-radius: 12px !important;
        border: none !important;
        font-size: 0.88rem !important;
    }}
    .stSuccess {{
        background: {success_bg} !important;
        color: {success_col} !important;
        border-left: 3px solid {success_col} !important;
    }}
    .stError {{
        background: {err_bg} !important;
        color: {err_col} !important;
        border-left: 3px solid {err_col} !important;
    }}

    /* ── 等級磚塊 ── */
    .metric-row {{
        display: flex; gap: 0.8rem; margin: 1rem 0 0.8rem;
    }}
    .metric-tile {{
        flex: 1;
        border-radius: 14px;
        padding: 1rem 0.6rem;
        text-align: center;
        border: 1px solid {border_col};
        transition: transform 0.18s, box-shadow 0.18s;
    }}
    .metric-tile:hover {{
        transform: translateY(-3px);
        box-shadow: 0 8px 28px {accent_glow};
    }}
    .metric-tile .grade-label {{
        font-size: 0.68rem;
        font-weight: 700;
        letter-spacing: 0.1em;
        text-transform: uppercase;
        margin-bottom: 0.35rem;
    }}
    .metric-tile .grade-count {{
        font-size: 2rem;
        font-weight: 800;
        line-height: 1;
    }}

    /* ── 平均列 ── */
    .avg-row {{
        display: flex; gap: 1.5rem; flex-wrap: wrap;
        background: {surface2};
        border: 1px solid {border_col};
        border-radius: 13px;
        padding: 0.9rem 1.3rem;
        margin-top: 0.8rem;
        font-size: 0.88rem;
        color: {text_sub};
    }}
    .avg-row strong {{
        background: {accent_grad};
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        font-weight: 700;
    }}
    .avg-sep {{ color: {border_col}; font-weight: 300; }}

    /* ── 門檻標籤 ── */
    .threshold-row {{ display: flex; gap: 0.55rem; flex-wrap: wrap; margin-top: 0.6rem; }}
    .threshold-badge {{
        font-size: 0.72rem;
        font-weight: 600;
        padding: 0.22rem 0.7rem;
        border-radius: 999px;
        background: {btn_bg};
        color: {accent_a};
        border: 1px solid {border_col};
        font-family: 'DM Mono', monospace;
    }}

    /* ── Section label（漸層）── */
    .section-label {{
        font-size: 0.69rem;
        font-weight: 700;
        letter-spacing: 0.13em;
        text-transform: uppercase;
        margin-bottom: 0.55rem;
        margin-top: 1.3rem;
        background: {accent_grad};
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        display: inline-block;
    }}

    /* ── Misc ── */
    .stMarkdown p {{ color: {text} !important; }}
    [data-testid="stCaptionContainer"] p {{ color: {text_sub} !important; }}
    hr {{ border-color: transparent !important; }}
    #MainMenu, footer, header {{ visibility: hidden !important; }}
    </style>
    """
    return css, toggle_icon, toggle_tip

dark = st.session_state.dark_mode
css_str, toggle_icon, toggle_tip = inject_css(dark)
st.markdown(css_str, unsafe_allow_html=True)

# ════════════════════════════════
#  樣式輔助 (原始功能，未修改)
# ════════════════════════════════
FONT_NAME = "新細明體"
FILLS = {
    "A++": None,
    "A+":  PatternFill("solid", fgColor="E6E6E6"),
    "A":   PatternFill("solid", fgColor="BFBFBF"),
    "B++": PatternFill("solid", fgColor="808080"),
    "":    PatternFill("solid", fgColor="808080"),
}

def _med(): return Side(style="medium", color="000000")
def _thn(): return Side(style="thin",   color="000000")
def all_thin():
    s = _thn()
    return Border(left=s, right=s, top=s, bottom=s)

def outer_med(r, c, r1, c1, r2, c2):
    return Border(
        left   = _med() if c == c1 else _thn(),
        right  = _med() if c == c2 else _thn(),
        top    = _med() if r == r1 else _thn(),
        bottom = _med() if r == r2 else _thn(),
    )

def sc(ws, row, col, value, bold=False, size=10, fill=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name=FONT_NAME, bold=bold, size=size)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if border: c.border = border
    if fill:   c.fill   = fill
    return c

def get_grade(total, th_app, th_ap, th_a, th_bpp):
    if total >= th_app: return "A++"
    if total >= th_ap:  return "A+"
    if total >= th_a:   return "A"
    if total >= th_bpp: return "B++"
    return ""

def read_students(uploaded):
    wb = openpyxl.load_workbook(uploaded, data_only=True)
    students = []
    for row in wb.active.iter_rows(values_only=True):
        if not row[0]: continue
        try:
            students.append((str(row[0]).strip(), float(row[1]), float(row[2]), float(row[3])))
        except (TypeError, ValueError): pass
    return students

# ════════════════════════════════
#  建立報表 (原始功能，未修改)
# ════════════════════════════════
def build_report(students, exam_lines, th_app, th_ap, th_a, th_bpp):
    sorted_s = sorted(students, key=lambda x: -x[3])
    n = len(sorted_s)

    avg_sel    = round(sum(s  for _, s,  _, _ in sorted_s) / n, 2)
    avg_nonsel = round(sum(ns for _, _, ns, _ in sorted_s) / n, 2)
    avg_total  = round(sum(t  for _, _, _,  t in sorted_s) / n, 2)
    
    counts = {"A++": 0, "A+": 0, "A": 0, "B++": 0}
    for _, _, _, t in sorted_s:
        g = get_grade(t, th_app, th_ap, th_a, th_bpp)
        if g in counts: counts[g] += 1

    rows_per_block = math.ceil((n + 1) / 3)
    remaining_space = rows_per_block - (n % rows_per_block)
    
    if (n % rows_per_block != 0) and remaining_space < 2:
        rows_per_block += 1
    elif (n % rows_per_block == 0):
        rows_per_block = rows_per_block

    HEADER_ROW = 1
    DATA_START = 2
    FINAL_ROW = DATA_START + rows_per_block - 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "成績報表"

    for b in range(3):
        base = b * 4 + 1
        ws.column_dimensions[get_column_letter(base)].width   = 9
        ws.column_dimensions[get_column_letter(base+1)].width = 7.5
        ws.column_dimensions[get_column_letter(base+2)].width = 7.5
        ws.column_dimensions[get_column_letter(base+3)].width = 7.5
    
    ws.column_dimensions["M"].width = 0.4
    ws.column_dimensions["N"].width = 7
    ws.column_dimensions["O"].width = 7
    ws.column_dimensions["P"].width = 7

    for b in range(3):
        base = b * 4 + 1
        for i, h in enumerate(["姓名", "選擇", "非選", "總分"]):
            sc(ws, HEADER_ROW, base + i, h, border=all_thin())

    for idx, (name, sel, nonsel, total) in enumerate(sorted_s):
        b = idx // rows_per_block
        r = DATA_START + (idx % rows_per_block)
        col = b * 4 + 1
        g = get_grade(total, th_app, th_ap, th_a, th_bpp)
        f = FILLS[g]
        sc(ws, r, col,     name,   fill=f, border=all_thin())
        sc(ws, r, col + 1, sel,    fill=f, border=all_thin())
        sc(ws, r, col + 2, nonsel, fill=f, border=all_thin())
        sc(ws, r, col + 3, total,  fill=f, border=all_thin())

    avg_pos = n
    b_avg = avg_pos // rows_per_block
    r_avg_start = DATA_START + (avg_pos % rows_per_block)
    r_avg_end = FINAL_ROW
    col_avg = b_avg * 4 + 1
    avg_vals = ["平均", avg_sel, avg_nonsel, avg_total]

    for i, val in enumerate(avg_vals):
        curr_col = col_avg + i
        for fill_r in range(r_avg_start, r_avg_end + 1):
            sc(ws, fill_r, curr_col, "", border=all_thin())
        if r_avg_end > r_avg_start:
            ws.merge_cells(start_row=r_avg_start, start_column=curr_col, end_row=r_avg_end, end_column=curr_col)
        sc(ws, r_avg_start, curr_col, val, bold=True, size=12, border=all_thin())

    for b in range(3):
        base = b * 4 + 1
        for r in range(DATA_START, FINAL_ROW + 1):
            if not ws.cell(row=r, column=base).border:
                for i in range(4):
                    sc(ws, r, base + i, "", border=all_thin())

    TITLE_R1, TITLE_R2, TITLE_C1, TITLE_C2 = 2, 7, 14, 16
    ws.merge_cells(start_row=TITLE_R1, start_column=TITLE_C1, end_row=TITLE_R2, end_column=TITLE_C2)
    tc = ws.cell(row=TITLE_R1, column=TITLE_C1)
    tc.value = "\n".join(exam_lines)
    tc.font = Font(name=FONT_NAME, bold=True, size=18)
    tc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(TITLE_R1, TITLE_R2 + 1):
        for c in range(TITLE_C1, TITLE_C2 + 1):
            ws.cell(row=r, column=c).border = outer_med(r, c, TITLE_R1, TITLE_C1, TITLE_R2, TITLE_C2)

    visible = [(g, counts[g]) for g in ["A++", "A+", "A", "B++"] if counts[g] > 0]
    GRADE_R1 = TITLE_R2 + 2
    for i, (g, cnt) in enumerate(visible):
        row = GRADE_R1 + i
        f = FILLS[g]
        for col, val in [(14, g), (15, cnt)]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(name=FONT_NAME, bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = outer_med(row, col, GRADE_R1, 14, GRADE_R1 + len(visible) - 1, 15)
            if f: cell.fill = f

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, counts, avg_sel, avg_nonsel, avg_total, n

# ════════════════════════════════
#  美化 UI
# ════════════════════════════════

# 模式切換按鈕（右上角）
toggle_col1, toggle_col2 = st.columns([8, 1])
with toggle_col2:
    if st.button(toggle_icon, help=toggle_tip, key="mode_btn"):
        st.session_state.dark_mode = not st.session_state.dark_mode
        st.rerun()

# 標題區
st.markdown("""
<div class="app-title">
  <span class="icon-badge">📊</span>
  <span class="title-text">成績報表產生器</span>
</div>
<div class="app-subtitle">上傳成績 Excel，自動排名並輸出格式化報表</div>
""", unsafe_allow_html=True)

# ── 考試設定 ──
with st.expander("⚙️　考試設定", expanded=True):
    exam_name = st.text_input(
        "考試名稱（以空格分三段）",
        placeholder="國三 金安模擬考 第六回",
        label_visibility="visible"
    )
    st.markdown('<div style="height:0.6rem"></div>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    th_app = c1.number_input("A++ 門檻", value=93.2, step=0.1, format="%.1f")
    th_ap  = c2.number_input("A+  門檻", value=85.7, step=0.1, format="%.1f")
    th_a   = c3.number_input("A   門檻", value=76.2, step=0.1, format="%.1f")
    th_bpp = c4.number_input("B++ 門檻", value=67.1, step=0.1, format="%.1f")

st.markdown('<hr class="styled-divider">', unsafe_allow_html=True)

# ── 上傳區 ──
st.markdown('<div class="section-label">📂 上傳成績檔案</div>', unsafe_allow_html=True)
uploaded = st.file_uploader("上傳 Excel 成績檔案（xlsx）", type=["xlsx", "xls"], label_visibility="collapsed")

if uploaded:
    try:
        students = read_students(uploaded)
        st.success(f"✅　成功讀取 **{len(students)}** 位學生資料")

        st.markdown('<div style="height:0.5rem"></div>', unsafe_allow_html=True)

        if st.button("🚀　產生報表", type="primary", use_container_width=True):
            if not exam_name.strip():
                st.error("⚠️　請先填入考試名稱")
            else:
                parts = exam_name.strip().split()
                if   len(parts) >= 3: lines = [parts[0], " ".join(parts[1:-1]), parts[-1]]
                elif len(parts) == 2: lines = [parts[0], "", parts[1]]
                else:                 lines = ["", exam_name.strip(), ""]

                buf, counts, avg_sel, avg_nonsel, avg_total, n = build_report(
                    students, lines, th_app, th_ap, th_a, th_bpp
                )

                st.markdown('<hr class="styled-divider">', unsafe_allow_html=True)

                # 下載按鈕
                st.download_button(
                    "⬇️　下載 Excel 報表",
                    data=buf,
                    file_name=f"{exam_name}.xlsx",
                    use_container_width=True,
                    type="primary"
                )

                # 等級磚塊（漸層）
                st.markdown('<div class="section-label" style="margin-top:1.25rem">📋 報表摘要</div>', unsafe_allow_html=True)

                if dark:
                    grade_styles = {
                        "A++": ("linear-gradient(135deg,rgba(79,142,247,0.25),rgba(167,139,250,0.20))", "#93c5fd", "rgba(79,142,247,0.3)"),
                        "A+":  ("linear-gradient(135deg,rgba(52,211,153,0.18),rgba(16,185,129,0.12))",  "#6ee7b7", "rgba(52,211,153,0.25)"),
                        "A":   ("linear-gradient(135deg,rgba(251,191,36,0.18),rgba(245,158,11,0.12))",  "#fcd34d", "rgba(251,191,36,0.25)"),
                        "B++": ("linear-gradient(135deg,rgba(248,113,113,0.18),rgba(239,68,68,0.12))",  "#fca5a5", "rgba(248,113,113,0.25)"),
                    }
                else:
                    grade_styles = {
                        "A++": ("linear-gradient(135deg,#dbeafe,#ede9fe)", "#2563eb", "rgba(59,110,240,0.15)"),
                        "A+":  ("linear-gradient(135deg,#d1fae5,#a7f3d0)",  "#065f46", "rgba(16,185,129,0.15)"),
                        "A":   ("linear-gradient(135deg,#fef9c3,#fde68a)",  "#92400e", "rgba(245,158,11,0.15)"),
                        "B++": ("linear-gradient(135deg,#fee2e2,#fecaca)",  "#991b1b", "rgba(239,68,68,0.15)"),
                    }
                tiles_html = '<div class="metric-row">'
                for g, (bg_tile, col_tile, border_tile) in grade_styles.items():
                    tiles_html += (
                        f'<div class="metric-tile" style="background:{bg_tile};border-color:{border_tile}">'
                        f'  <div class="grade-label" style="color:{col_tile}">{g}</div>'
                        f'  <div class="grade-count" style="color:{col_tile}">{counts[g]}</div>'
                        f'</div>'
                    )
                tiles_html += '</div>'
                st.markdown(tiles_html, unsafe_allow_html=True)

                # 平均列
                st.markdown(
                    f'<div class="avg-row">'
                    f'<span>📐 選擇題均分　<strong>{avg_sel}</strong></span>'
                    f'<span class="avg-sep">|</span>'
                    f'<span>非選均分　<strong>{avg_nonsel}</strong></span>'
                    f'<span class="avg-sep">|</span>'
                    f'<span>總分均分　<strong>{avg_total}</strong></span>'
                    f'<span class="avg-sep">|</span>'
                    f'<span>共　<strong>{n}</strong> 人</span>'
                    f'</div>',
                    unsafe_allow_html=True
                )

    except Exception as e:
        st.error(f"❌　讀取失敗：{e}")
